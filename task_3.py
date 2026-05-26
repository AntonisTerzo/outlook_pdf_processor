import re
from pathlib import Path
from collections import defaultdict
from config import TASK3_FILE_TYPES
from pdf_utils import (
    detect_file_type_task3,
    extract_combo_id_from_filename,
    extract_packing_list_data,
    extract_hs_codes,
    extract_cargo_description_from_subject,
)
from outlook_utils import (
    connect_to_outlook, find_outlook_folder,
    initialize_com, uninitialize_com,
    create_unique_folder
)
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill


def clean_email_subject(subject):
    """Strip whitespace and replace Windows-invalid characters with underscore."""
    cleaned = (subject or "").strip()
    invalid_chars = r'[<>:"/\\|?*.]'
    cleaned = re.sub(invalid_chars, '_', cleaned)
    return cleaned


def _save_attachments_from_message(message, outlook, dest_folder, log_func):
    """
    Save all PDF attachments from an email (including PDFs inside .msg attachments)
    into dest_folder. Returns list of saved Path objects.
    """
    saved = []

    if not (hasattr(message, 'Attachments') and message.Attachments.Count > 0):
        return saved

    for attachment in message.Attachments:
        filename = attachment.FileName

        if filename.lower().endswith('.pdf'):
            target = dest_folder / filename
            # Avoid overwrite collisions
            counter = 1
            while target.exists():
                stem = filename[:-4] if filename.lower().endswith('.pdf') else filename
                target = dest_folder / f"{stem}({counter}).pdf"
                counter += 1
            attachment.SaveAsFile(str(target))
            saved.append(target)
            log_func(f"  Saved PDF: {target.name}")

        elif filename.lower().endswith('.msg'):
            log_func(f"  Opening .msg file: {filename}")
            temp_msg_path = dest_folder / filename
            attachment.SaveAsFile(str(temp_msg_path))
            try:
                namespace = outlook.GetNamespace("MAPI")
                msg = namespace.OpenSharedItem(str(temp_msg_path))
                if msg.Attachments.Count > 0:
                    for inner in msg.Attachments:
                        if inner.FileName.lower().endswith('.pdf'):
                            target = dest_folder / inner.FileName
                            counter = 1
                            while target.exists():
                                stem = inner.FileName[:-4]
                                target = dest_folder / f"{stem}({counter}).pdf"
                                counter += 1
                            inner.SaveAsFile(str(target))
                            saved.append(target)
                            log_func(f"    Saved PDF from .msg: {target.name}")
            except Exception as e:
                log_func(f"    Error opening .msg: {e}")
            finally:
                try:
                    temp_msg_path.unlink()
                except Exception:
                    pass

    return saved


def _group_combos(pdf_paths, log_func):
    """
    Group PDFs by their combo identifier (longest digit run in filename),
    and within each combo detect their file types.
    Returns dict: {combo_id: {"INVOICE": Path|None, "PACKING_LIST": Path|None,
                              "CUSTOMS_CODE": Path|None, "unknown": [Path,...]}}
    """
    combos = defaultdict(lambda: {
        "INVOICE": None,
        "PACKING_LIST": None,
        "CUSTOMS_CODE": None,
        "unknown": [],
    })

    for pdf_path in pdf_paths:
        combo_id = extract_combo_id_from_filename(pdf_path.name)
        if not combo_id:
            log_func(f"  ! Could not extract combo id from: {pdf_path.name}")
            combos["NO_ID"]["unknown"].append(pdf_path)
            continue

        file_type = detect_file_type_task3(pdf_path, TASK3_FILE_TYPES)
        if file_type is None:
            log_func(f"  ! Could not detect type of: {pdf_path.name}")
            combos[combo_id]["unknown"].append(pdf_path)
        else:
            # If we already have one of this type, keep the first and log conflict
            if combos[combo_id][file_type] is None:
                combos[combo_id][file_type] = pdf_path
                log_func(f"  {pdf_path.name} -> combo {combo_id} / {file_type}")
            else:
                log_func(
                    f"  ! Duplicate {file_type} in combo {combo_id}: "
                    f"{pdf_path.name} (keeping {combos[combo_id][file_type].name})"
                )
                combos[combo_id]["unknown"].append(pdf_path)

    return combos


def _build_combo_row(combo_id, combo_files, mail_description, log_func):
    """
    Build the 9-column row for a combo.
    Returns dict with keys matching column headers, plus '_notes' list.
    Returns None if packing list or customs code is missing.
    """
    notes = []

    packing = combo_files.get("PACKING_LIST")
    customs = combo_files.get("CUSTOMS_CODE")
    invoice = combo_files.get("INVOICE")

    if packing is None:
        notes.append("missing PACKING_LIST")
    if customs is None:
        notes.append("missing CUSTOMS_CODE")
    if invoice is None:
        notes.append("missing INVOICE (continuing - invoice is optional)")

    # Per spec: process if invoice is missing, otherwise notify the user.
    # So packing_list and customs_code are required.
    if packing is None or customs is None:
        return None, notes

    pl_data = extract_packing_list_data(packing)
    hs_codes = extract_hs_codes(customs)

    row = {
        "Cargo Description (Mail)": mail_description or "",
        "Cargo Description": ", ".join(pl_data["descriptions"]) if pl_data["descriptions"] else "",
        "IV": pl_data["iv"] or "",
        "SRN": pl_data["srn"] or "",
        "PCS": pl_data["pcs"] or "",
        "KG": pl_data["kg"] or "",
        "M3": pl_data["m3"] or "",
        "DIMS": pl_data["dims"] or "",
        "HS Code": ", ".join(hs_codes) if hs_codes else "",
        "_combo_id": combo_id,
    }
    return row, notes


def _create_excel_report(rows, output_folder, log_func):
    """Create the consolidated Excel report - one sheet, one row per combo."""
    if not rows:
        return None

    excel_path = output_folder / "Task_3_Report.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Combos"

    headers = [
        "Cargo Description (Mail)",
        "Cargo Description",
        "IV",
        "SRN",
        "PCS",
        "KG",
        "M3",
        "DIMS",
        "HS Code",
    ]

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=row_idx, column=col_idx, value=row.get(header, ""))

    # Column widths
    widths = {"A": 25, "B": 40, "C": 18, "D": 18, "E": 10, "F": 12, "G": 12, "H": 18, "I": 30}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    # Freeze header row
    ws.freeze_panes = "A2"

    wb.save(excel_path)
    return excel_path


def run_task_3(log_func=print):
    """
    Task 3:
    - Connect to Outlook, find Inbox/Task_3 folder.
    - Save all PDF attachments (including PDFs inside .msg attachments).
    - Group PDFs into combos by the longest digit-run in their filename.
    - Each combo should contain: INVOICE, PACKING_LIST, CUSTOMS_CODE.
      Process if INVOICE is missing; require PACKING_LIST and CUSTOMS_CODE.
    - Build a 9-column Excel report (one row per combo).
    Returns: (processed_count, incomplete_combos_list, folder_path)
    """
    initialize_com()

    try:
        log_func("Task 3: Attempting to connect to Outlook...")
        outlook, namespace = connect_to_outlook()
        log_func("Successfully connected to Outlook")

        task_folder = find_outlook_folder(namespace, "Task_3")
        if not task_folder:
            log_func("Error: Task_3 folder not found in Inbox")
            return 0, [], None

        log_func(f"Found folder: {task_folder.Name}")

        if task_folder.Items.Count == 0:
            log_func("\nThere were no emails to process inside Task_3 folder.")
            return 0, [], None

        log_func(f"Processing {task_folder.Items.Count} emails\n")

        downloads = Path.home() / "Downloads"
        pdf_folder = create_unique_folder(downloads, "outlook_pdf_processor_task_3")
        log_func(f"Created folder: {pdf_folder.name}")

        temp_folder = pdf_folder / "temp"
        temp_folder.mkdir(exist_ok=True)

        all_rows = []
        incomplete_combos = []  # list of (email_subject, combo_id, notes)
        total_processed = 0

        for message in task_folder.Items:
            if not (hasattr(message, 'Attachments') and message.Attachments.Count > 0):
                continue

            email_subject_raw = message.Subject if message.Subject else "No_Subject"
            email_subject = clean_email_subject(email_subject_raw)[:300]
            mail_description = extract_cargo_description_from_subject(email_subject_raw)

            log_func(f"\nProcessing email: {email_subject}")
            if mail_description:
                log_func(f"  Cargo Description (Mail): {mail_description}")
            else:
                log_func("  ! Could not find word before 'CHINA' in subject")

            # Each email gets its own subfolder for the PDFs
            base_email_folder = pdf_folder / email_subject
            email_folder = base_email_folder
            counter = 1
            while email_folder.exists():
                email_folder = pdf_folder / f"{email_subject}({counter})"
                counter += 1
            email_folder.mkdir(exist_ok=True)

            # Save all PDFs
            saved_pdfs = _save_attachments_from_message(
                message, outlook, email_folder, log_func
            )

            if not saved_pdfs:
                log_func("  No PDF attachments found in this email")
                continue

            # Group into combos
            combos = _group_combos(saved_pdfs, log_func)

            # Build rows
            for combo_id, files in combos.items():
                if combo_id == "NO_ID":
                    for f in files["unknown"]:
                        incomplete_combos.append(
                            (email_subject_raw, "no-id", [f"unidentified file: {f.name}"])
                        )
                    continue

                row, notes = _build_combo_row(combo_id, files, mail_description, log_func)
                if row is None:
                    log_func(f"  ! Combo {combo_id} incomplete: {', '.join(notes)}")
                    incomplete_combos.append((email_subject_raw, combo_id, notes))
                else:
                    if any("missing INVOICE" in n for n in notes):
                        log_func(f"  Combo {combo_id}: processed without INVOICE")
                    else:
                        log_func(f"  Combo {combo_id}: complete")
                    all_rows.append(row)
                    total_processed += 1

        # Create the consolidated Excel
        if all_rows:
            log_func("\nCreating Excel report...")
            excel_path = _create_excel_report(all_rows, pdf_folder, log_func)
            if excel_path:
                log_func(f"Excel report created: {excel_path.name}")

        log_func("\n" + "=" * 60)
        log_func("Task 3 Processing complete")
        log_func(f"{total_processed} combos written to Excel")
        log_func(f"Location: {pdf_folder}")
        if incomplete_combos:
            log_func(f"\nWARNING: {len(incomplete_combos)} incomplete combo(s):")
            for subj, cid, notes in incomplete_combos:
                log_func(f"  - {subj} / combo {cid}: {', '.join(notes)}")
        log_func("=" * 60)

        # Clean up temp folder if empty
        if temp_folder.exists():
            remaining = list(temp_folder.glob('*'))
            if not remaining:
                temp_folder.rmdir()

        return total_processed, incomplete_combos, pdf_folder

    except Exception as e:
        log_func(f"\nError: {e}")
        import traceback
        traceback.print_exc()
        return 0, [], None

    finally:
        uninitialize_com()
