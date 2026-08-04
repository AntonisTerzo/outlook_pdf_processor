import re
from pathlib import Path
from collections import defaultdict
from config import TASK3_FILE_TYPES
from pdf_utils import (
    detect_file_type_task3,
    extract_combo_id_from_filename,
    extract_packing_list_data,
    extract_hs_codes,
    extract_cargo_description_from_subject,
    set_extraction_logger,
)
from outlook_utils import (
    connect_to_outlook, find_outlook_folder,
    initialize_com, uninitialize_com,
    create_unique_folder
)
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill


def _read_attachment_bytes(attachment):
    """
    Read an Outlook attachment's content into memory as bytes WITHOUT saving
    a permanent copy. Outlook's COM API has no direct in-memory read, so we
    write to a short-lived temp file, read it back, and delete it immediately.
    """
    import tempfile
    import os

    tmp_path = None
    try:
        fd, tmp_path = tempfile.mkstemp(suffix=".pdf")
        os.close(fd)
        attachment.SaveAsFile(tmp_path)
        with open(tmp_path, "rb") as f:
            return f.read()
    except Exception as e:
        print(f"Error reading attachment bytes: {e}")
        return None
    finally:
        if tmp_path:
            try:
                os.remove(tmp_path)
            except Exception:
                pass


def _collect_pdfs_from_message(message, outlook, log_func):
    """
    Walk an email's attachments and return a list of (filename, pdf_bytes).
    PDFs nested inside .msg attachments are also collected. Nothing is saved
    to a permanent location.
    """
    collected = []

    if not (hasattr(message, 'Attachments') and message.Attachments.Count > 0):
        return collected

    for attachment in message.Attachments:
        filename = attachment.FileName

        if filename.lower().endswith('.pdf'):
            data = _read_attachment_bytes(attachment)
            if data:
                collected.append((filename, data))
                log_func(f"  Read PDF: {filename}")

        elif filename.lower().endswith('.msg'):
            log_func(f"  Opening .msg file: {filename}")
            import tempfile
            import os
            tmp_msg = None
            try:
                fd, tmp_msg = tempfile.mkstemp(suffix=".msg")
                os.close(fd)
                attachment.SaveAsFile(tmp_msg)
                namespace = outlook.GetNamespace("MAPI")
                msg = namespace.OpenSharedItem(tmp_msg)
                if msg.Attachments.Count > 0:
                    for inner in msg.Attachments:
                        if inner.FileName.lower().endswith('.pdf'):
                            data = _read_attachment_bytes(inner)
                            if data:
                                collected.append((inner.FileName, data))
                                log_func(f"    Read PDF from .msg: {inner.FileName}")
            except Exception as e:
                log_func(f"    Error opening .msg: {e}")
            finally:
                if tmp_msg:
                    try:
                        os.remove(tmp_msg)
                    except Exception:
                        pass

    return collected


def _group_combos(pdfs, log_func):
    """
    Group (filename, bytes) PDFs by combo id (longest digit run in filename),
    detecting each file's type. The invoice file maps to None and is skipped.

    Returns (combos, used_files):
        combos    -> {combo_id: {"PACKING_LIST": bytes|None,
                                 "CUSTOMS_CODE": bytes|None}}
        used_files-> [(filename, bytes)] for the files actually used, i.e. only
                     the PACKING_LIST / CUSTOMS_CODE pairs we read. These are
                     the files that get saved to disk; invoices, unrecognised
                     files and dropped duplicates are not included.
    """
    combos = defaultdict(lambda: {"PACKING_LIST": None, "CUSTOMS_CODE": None})
    used_files = []

    for filename, data in pdfs:
        combo_id = extract_combo_id_from_filename(filename)
        if not combo_id:
            log_func(f"  ! Could not extract combo id from: {filename}")
            continue

        file_type = detect_file_type_task3(data, TASK3_FILE_TYPES)
        if file_type is None:
            # Most likely the invoice file, which we intentionally ignore.
            log_func(f"  (skipping non-target file: {filename})")
            continue

        if combos[combo_id][file_type] is None:
            combos[combo_id][file_type] = data
            used_files.append((filename, data))
            log_func(f"  {filename} -> combo {combo_id} / {file_type}")
        else:
            log_func(f"  ! Duplicate {file_type} in combo {combo_id}: {filename} (keeping first)")

    return combos, used_files


def _sanitize_folder_name(subject):
    """
    Turn an email subject into a safe Windows folder name: strip characters
    Windows forbids, collapse whitespace and cap the length.
    """
    name = re.sub(r'[<>:"/\\|?*]', '', subject or "")
    name = re.sub(r'\s+', ' ', name).strip()
    # Windows also refuses names ending in a dot or space.
    name = name.rstrip('. ')
    if len(name) > 100:
        name = name[:100].rstrip('. ')
    return name or "No_Subject"


def _save_email_pdfs(output_folder, subject, used_files, log_func):
    """
    Save the PDFs we actually read for one email into its own folder inside
    output_folder. Returns the folder path, or None if nothing was saved.
    """
    if not used_files:
        return None

    email_folder = create_unique_folder(
        output_folder, _sanitize_folder_name(subject))

    saved = 0
    seen_names = {}
    for filename, data in used_files:
        target = Path(filename).name
        # Two attachments in the same email can share a name (e.g. one nested
        # in a .msg): keep both by suffixing the later ones.
        if target in seen_names:
            seen_names[target] += 1
            stem = Path(target).stem
            suffix = Path(target).suffix
            target = f"{stem}_{seen_names[target]}{suffix}"
        else:
            seen_names[target] = 0

        try:
            with open(email_folder / target, "wb") as f:
                f.write(data)
            saved += 1
        except Exception as e:
            log_func(f"  ! Could not save {target}: {e}")

    log_func(f"  Saved {saved} PDF(s) to: {email_folder}")
    return email_folder


def _build_combo_row(combo_id, combo_files, mail_description, log_func):
    """
    Build the 9-column row for a combo.

    PACKING_LIST is required. CUSTOMS_CODE is NOT required: if it is missing
    but the packing list is present, the row is still produced with every
    packing-list field filled in, and only the HS Code cell carries the
    "NOT FOUND - check PDF" placeholder (that single cell is shaded red).

    For any packing-list field that cannot be extracted (description, IV, SRN,
    PCS, KG, M3, DIMS), the cell gets the placeholder AND the ENTIRE row is
    shaded red so the user can spot it at a glance.

    Returns (row_dict, notes_list, missing_fields_list). row is None only when
    the PACKING_LIST itself is missing (nothing to build from).
    """
    notes = []
    missing_fields = []
    packing = combo_files.get("PACKING_LIST")
    customs = combo_files.get("CUSTOMS_CODE")

    if packing is None:
        notes.append("missing PACKING_LIST")
        # Without a packing list there is nothing to build the row from.
        return None, notes, missing_fields

    pl = extract_packing_list_data(packing)

    placeholder = "NOT FOUND - check PDF"

    # Tracks which fields triggered a whole-row warning (packing-list misses).
    row_red = False

    def field(value, name, *, is_list=False, joiner=", "):
        """Return the cell value, or a placeholder if missing; log misses and
        flag the row for red shading (packing-list fields only)."""
        nonlocal row_red
        if is_list:
            if value:
                return joiner.join(value)
        else:
            if value not in (None, ""):
                return value
        missing_fields.append(name)
        row_red = True
        log_func(f"    [WARNING] combo {combo_id}: could not extract {name}")
        return placeholder

    # HS Code comes from the CUSTOMS_CODE file. If that file is absent we still
    # write the row, but the HS cell alone is marked (not the whole row).
    hs_only_red = False
    if customs is None:
        notes.append("missing CUSTOMS_CODE")
        hs_value = placeholder
        hs_only_red = True
        missing_fields.append("HS Code (no customs file)")
        log_func(f"    [WARNING] combo {combo_id}: no CUSTOMS_CODE file - HS Code not available")
    else:
        hs_codes = extract_hs_codes(customs)
        if hs_codes:
            hs_value = ";".join(hs_codes)
        else:
            # Customs file present but no codes found inside -> treat like a
            # missing packing-list field: placeholder + whole-row red.
            hs_value = placeholder
            missing_fields.append("HS Code")
            row_red = True
            log_func(f"    [WARNING] combo {combo_id}: could not extract HS Code")

    row = {
        "Cargo Description (Mail)": mail_description or placeholder,
        "Cargo Description": field(pl["descriptions"], "Cargo Description", is_list=True),
        "IV": field(pl["iv"], "IV"),
        "SRN": field(pl["srn"], "SRN"),
        "PCS": field(pl["pcs"], "PCS"),
        "KG": field(pl["kg"], "KG"),
        "M3": field(pl["m3"], "M3"),
        "DIMS": field(pl["dims"], "DIMS"),
        "HS Code": hs_value,
    }

    if not mail_description:
        missing_fields.append("Cargo Description (Mail)")
        row_red = True
        log_func(f"    [WARNING] combo {combo_id}: could not extract Cargo Description (Mail) from subject")

    # Colour instructions for the Excel writer:
    #   row["_row_red"]    -> shade the whole row red
    #   row["_hs_cell_red"]-> shade only the HS Code cell red
    # (whole-row red supersedes the single-cell case)
    row["_row_red"] = row_red
    row["_hs_cell_red"] = hs_only_red and not row_red

    if missing_fields:
        log_func(f"  ! Combo {combo_id}: {len(missing_fields)} field(s) not found: {', '.join(missing_fields)}")

    return row, notes, missing_fields


def _create_excel_report(rows, output_folder):
    """Create the consolidated Excel report - one sheet, one row per combo."""
    if not rows:
        return None

    excel_path = output_folder / "Task_3_Report.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Combos"

    headers = [
        "Cargo Description (Mail)",
        "Cargo Description",
        "IV",
        "SRN",
        "PCS",
        "KG",
        "M3",
        "DIMS",
        "HS Code",
    ]

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    # Red shading for missing data.
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    hs_col_idx = headers.index("HS Code") + 1  # 1-based column of HS Code

    for row_idx, row in enumerate(rows, start=2):
        row_red = row.get("_row_red", False)
        hs_cell_red = row.get("_hs_cell_red", False)
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row.get(header, ""))
            if row_red:
                # A packing-list field was missing: shade the entire row.
                cell.fill = red_fill
            elif hs_cell_red and col_idx == hs_col_idx:
                # Only the HS Code is missing (no customs file): shade just it.
                cell.fill = red_fill

    widths = {"A": 25, "B": 40, "C": 18, "D": 18, "E": 10, "F": 12, "G": 12, "H": 18, "I": 30}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ws.freeze_panes = "A2"
    wb.save(excel_path)
    return excel_path


def run_task_3(log_func=print):
    """
    Task 3:
    - Connect to Outlook, find Inbox/Task_3 folder (notify if missing).
    - Read all PDF attachments into memory.
    - Group PDFs into combos by the longest digit-run in their filename.
    - Each combo needs PACKING_LIST + CUSTOMS_CODE (invoice is ignored).
    - Save the PACKING_LIST / CUSTOMS_CODE PDFs we read into a per-email
      subfolder of the output folder (other attachments are not saved).
    - Extract data and write one consolidated Task_3_Report.xlsx to Downloads.
    Returns: (processed_count, incomplete_combos, combos_with_missing_fields,
              output_folder_path)
    """
    initialize_com()
    # Route low-level extraction diagnostics (ambiguous/missing values) into
    # the same log the user sees. Restored in the finally block.
    _previous_logger = set_extraction_logger(log_func)

    try:
        log_func("Task 3: Attempting to connect to Outlook...")
        outlook, namespace = connect_to_outlook()
        log_func("Successfully connected to Outlook")

        task_folder = find_outlook_folder(namespace, "Task_3")
        if not task_folder:
            log_func("Error: Task_3 folder not found in Inbox")
            return 0, [], [], None

        log_func(f"Found folder: {task_folder.Name}")

        if task_folder.Items.Count == 0:
            log_func("\nThere were no emails to process inside Task_3 folder.")
            return 0, [], [], None

        log_func(f"Processing {task_folder.Items.Count} emails\n")

        all_rows = []
        incomplete_combos = []  # list of (email_subject, combo_id, notes)
        combos_with_missing_fields = []  # list of (email_subject, combo_id, missing_fields)
        total_processed = 0

        # The output folder in Downloads holds one subfolder per email plus the
        # final Excel. It is created lazily so nothing is left behind when
        # there is nothing to save.
        output_folder = None

        def get_output_folder():
            nonlocal output_folder
            if output_folder is None:
                downloads = Path.home() / "Downloads"
                output_folder = create_unique_folder(
                    downloads, "outlook_pdf_processor_task_3")
            return output_folder

        for message in task_folder.Items:
            if not (hasattr(message, 'Attachments') and message.Attachments.Count > 0):
                continue

            subject_raw = message.Subject if message.Subject else "No_Subject"
            mail_description = extract_cargo_description_from_subject(subject_raw)

            log_func(f"\nProcessing email: {subject_raw}")
            if mail_description:
                log_func(f"  Cargo Description (Mail): {mail_description}")
            else:
                log_func("  ! Could not find description (- word - CHINA) in subject")

            pdfs = _collect_pdfs_from_message(message, outlook, log_func)
            if not pdfs:
                log_func("  No PDF attachments found in this email")
                continue

            combos, used_files = _group_combos(pdfs, log_func)

            # Save only the pairs we read (packing list / customs code).
            if used_files:
                _save_email_pdfs(get_output_folder(), subject_raw,
                                 used_files, log_func)

            for combo_id, files in combos.items():
                row, notes, missing_fields = _build_combo_row(
                    combo_id, files, mail_description, log_func)
                if row is None:
                    log_func(f"  ! Combo {combo_id} incomplete: {', '.join(notes)}")
                    incomplete_combos.append((subject_raw, combo_id, notes))
                else:
                    if missing_fields:
                        log_func(f"  Combo {combo_id}: written with {len(missing_fields)} missing field(s)")
                        combos_with_missing_fields.append((subject_raw, combo_id, missing_fields))
                    else:
                        log_func(f"  Combo {combo_id}: complete")
                    all_rows.append(row)
                    total_processed += 1

        # Write the consolidated Excel next to the saved email folders
        if all_rows:
            log_func("\nCreating Excel report...")
            excel_path = _create_excel_report(all_rows, get_output_folder())
            if excel_path:
                log_func(f"Excel report created: {excel_path}")

        log_func("\n" + "=" * 60)
        log_func("Task 3 Processing complete")
        log_func(f"{total_processed} combo(s) written to Excel")
        if output_folder:
            log_func(f"Location: {output_folder}")
        if incomplete_combos:
            log_func(f"\nWARNING: {len(incomplete_combos)} incomplete combo(s) (skipped):")
            for subj, cid, notes in incomplete_combos:
                log_func(f"  - {subj} / combo {cid}: {', '.join(notes)}")
        if combos_with_missing_fields:
            log_func(f"\nWARNING: {len(combos_with_missing_fields)} combo(s) written with missing fields:")
            for subj, cid, fields in combos_with_missing_fields:
                log_func(f"  - {subj} / combo {cid}: {', '.join(fields)}")
        log_func("=" * 60)

        return total_processed, incomplete_combos, combos_with_missing_fields, output_folder

    except Exception as e:
        log_func(f"\nError: {e}")
        import traceback
        traceback.print_exc()
        return 0, [], [], None

    finally:
        set_extraction_logger(_previous_logger)
        uninitialize_com()
