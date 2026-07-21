import re
from pathlib import Path
from collections import defaultdict, Counter
from config import TASK2_CITIES, BRASILIEN_SUBCITIES, LYMAN_ELECTRONICS_SUBCITIES, MANUAL_REVIEW_FALLBACK_CITIES
from pdf_utils import (
    extract_city_from_filename_task2, extract_dimensions_from_pdf,
    check_city_inside_pdf, check_variofix_in_pdf
)
from outlook_utils import (
    connect_to_outlook, find_outlook_folder,
    process_msg_file, initialize_com, uninitialize_com,
    create_unique_folder
)
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill


def clean_email_subject(subject):
    """
    Clean email subject to be used as folder name.
    Steps:
    1. Strip leading and trailing whitespaces
    2. Replace Windows invalid characters with underscores
       Invalid characters: < > : " / \ | ? *
    """

    cleaned = subject.strip()

    # Replace all Windows invalid characters with underscore
    invalid_chars = r'[<>:"/\\|?*.]'
    cleaned = re.sub(invalid_chars, '_', cleaned)

    return cleaned


# Track city counts for numbering
city_counter = defaultdict(int)

# Store dimensions data: {city: {file_number: [dimensions]}}
dimensions_data = defaultdict(lambda: defaultdict(list))


def process_pdf_task2(temp_pdf_path, output_folder, original_filename):
    """
    Process a single PDF for Task 2.
    Extracts city from filename, extracts dimensions, and renames with numbering.
    Special handling:
    - BRASILIEN (not JOINVILLE): Check inside PDF for INDAIATUBA or RIO CLARO
    - LYMAN ELECTRONICS: Check inside PDF for MAXOLUTION
    - Manual review files: Check inside PDF for TAPUKARA
    Returns dict with 'success' and 'message' keys.
    """
    city = extract_city_from_filename_task2(original_filename, TASK2_CITIES)

    if city:
        # Special handling for BRASILIEN (not BRASILIEN JOINVILLE)
        if city == "BRASILIEN":
            subcity = check_city_inside_pdf(temp_pdf_path, BRASILIEN_SUBCITIES)
            if subcity:
                city = f"BRASILIEN {subcity}"

        # Special handling for LYMAN ELECTRONICS
        if city == "LYMAN ELECTRONICS":
            subcity = check_city_inside_pdf(
                temp_pdf_path, LYMAN_ELECTRONICS_SUBCITIES)
            if subcity:
                city = subcity

        # Increment counter for this city
        city_counter[city] += 1
        count = city_counter[city]

        # Extract dimensions from PDF
        dimensions, warning = extract_dimensions_from_pdf(temp_pdf_path)

        # Store dimensions for this file
        dimensions_data[city][count] = dimensions

        # Create filename: "BRASILIEN INDAIATUBA 1.pdf", "MAXOLUTION 1.pdf", etc.
        # If the PDF mentions "variofix", prefix the name so the file is easy
        # to spot in the folder and the user is alerted in the messagebox.
        has_variofix = check_variofix_in_pdf(temp_pdf_path)
        new_filename = f"{city} {count}.pdf"
        if has_variofix:
            new_filename = f"VARIOFIX_{new_filename}"
        final_path = output_folder / new_filename
        temp_pdf_path.rename(final_path)

        # Build message with warning if needed
        message = f"Saved as: {new_filename} ({len(dimensions)} dimensions found)"
        if warning:
            message += f" - {warning}"

        return {
            'success': True,
            'message': message,
            'variofix': has_variofix,
            'filename': new_filename,
        }
    else:
        # Manual review - check for fallback cities inside PDF
        fallback_city = check_city_inside_pdf(
            temp_pdf_path, MANUAL_REVIEW_FALLBACK_CITIES)

        if fallback_city:
            # Found a fallback city (e.g., TAPUKARA)
            city_counter[fallback_city] += 1
            count = city_counter[fallback_city]

            # Extract dimensions
            dimensions, warning = extract_dimensions_from_pdf(temp_pdf_path)
            dimensions_data[fallback_city][count] = dimensions

            has_variofix = check_variofix_in_pdf(temp_pdf_path)
            new_filename = f"{fallback_city} {count}.pdf"
            if has_variofix:
                new_filename = f"VARIOFIX_{new_filename}"
            final_path = output_folder / new_filename
            temp_pdf_path.rename(final_path)

            # Build message with warning if needed
            message = f"Saved as: {new_filename} (found via PDF scan)"
            if warning:
                message += f" - {warning}"

            return {
                'success': True,
                'message': message,
                'variofix': has_variofix,
                'filename': new_filename,
            }
        else:
            # Move to MANUAL REVIEW folder - but still extract dimensions
            manual_folder = output_folder / "MANUAL REVIEW"
            manual_folder.mkdir(exist_ok=True)

            # Extract dimensions even for manual review files
            dimensions, warning = extract_dimensions_from_pdf(temp_pdf_path)

            # Store dimensions under "MANUAL REVIEW" category with original filename
            if dimensions:
                dimensions_data["MANUAL REVIEW"][original_filename] = dimensions

            # If variofix appears in the PDF, prefix the base filename before
            # applying collision-avoidance numbering.
            has_variofix = check_variofix_in_pdf(temp_pdf_path)
            base_name = original_filename
            if has_variofix:
                base_name = f"VARIOFIX_{original_filename}"

            # Create unique filename if file already exists
            counter = 1
            final_path = manual_folder / base_name

            while final_path.exists():
                name_without_ext = base_name.replace(
                    '.pdf', '').replace('.PDF', '')
                final_path = manual_folder / \
                    f"{name_without_ext}({counter}).pdf"
                counter += 1

            temp_pdf_path.rename(final_path)

            # Build message with dimension count and warning
            dim_msg = f" ({len(dimensions)} dimensions found)" if dimensions else ""
            if warning:
                dim_msg += f" - {warning}"

            return {
                'success': False,
                'message': f"Moved to MANUAL REVIEW: {final_path.name}{dim_msg}",
                'variofix': has_variofix,
                'filename': final_path.name,
            }


def create_excel_report(output_folder):
    """
    Create Excel file with dimensions grouped by city.
    Each city gets its own sheet with files separated.
    """
    if not dimensions_data:
        return None

    excel_path = output_folder / "Dimensions_Report.xlsx"
    wb = openpyxl.Workbook()

    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    # Create a sheet for each city
    for city in sorted(dimensions_data.keys()):
        ws = wb.create_sheet(title=city[:31])  # Excel sheet names max 31 chars

        # Header styling
        header_fill = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        # Add headers
        ws['A1'] = 'File'
        ws['B1'] = 'Dimension'
        ws['C1'] = 'Count'

        for cell in ['A1', 'B1', 'C1']:
            ws[cell].fill = header_fill
            ws[cell].font = header_font
            ws[cell].alignment = Alignment(
                horizontal='center', vertical='center')

        row = 2

        # Process each file for this city
        for file_num in sorted(dimensions_data[city].keys()):
            dimensions_list = dimensions_data[city][file_num]

            if not dimensions_list:
                # No dimensions found
                ws[f'A{row}'] = f"{city} {file_num}"
                ws[f'B{row}'] = "No dimensions found"
                ws[f'C{row}'] = 0
                row += 1
            else:
                # Group and count dimensions
                dimension_counts = Counter(dimensions_list)

                # Sort by dimension for consistency
                for dimension, count in sorted(dimension_counts.items()):
                    ws[f'A{row}'] = f"{city} {file_num}"
                    ws[f'B{row}'] = dimension
                    ws[f'C{row}'] = count
                    row += 1

            # Add separator row
            row += 1

        # Auto-adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 10

    wb.save(excel_path)
    return excel_path


def run_task_2(log_func=print):
    """
    Main Task 2 logic: Download PDFs from Task_2 folder, rename by city with numbering,
    extract dimensions and create Excel report.
    Each email gets its own subfolder named after the email subject.
    """
    # Reset counters and data for this run
    city_counter.clear()
    dimensions_data.clear()

    initialize_com()

    try:
        log_func("Task 2: Attempting to connect to Outlook...")
        outlook, namespace = connect_to_outlook()
        log_func("Successfully connected to Outlook")

        # Find Task_2 folder
        task_folder = find_outlook_folder(namespace, "Task_2")

        if not task_folder:
            log_func("Error: Task_2 folder not found in Inbox")
            return 0, 0, None, [], []

        log_func(f"Found folder: {task_folder.Name}")

        # Check if there are any emails to process
        if task_folder.Items.Count == 0:
            log_func("\nThere were no emails to process inside Task_2 folder.")
            return 0, 0, None, [], []

        log_func(f"Processing {task_folder.Items.Count} emails\n")

        # Only create folders if there are emails to process
        downloads = Path.home() / "Downloads"
        pdf_folder = create_unique_folder(
            downloads, "outlook_pdf_processor_task_2")
        log_func(f"Created folder: {pdf_folder.name}")

        temp_folder = pdf_folder / "temp"
        temp_folder.mkdir(exist_ok=True)

        total_processed_count = 0
        total_manual_review_count = 0
        files_with_warnings = []
        variofix_files = []  # final filenames of any PDFs containing "variofix"

        # Process each email - each gets its own subfolder
        for message in task_folder.Items:
            if hasattr(message, 'Attachments') and message.Attachments.Count > 0:
                email_subject = message.Subject if message.Subject else "No_Subject"
                email_subject = clean_email_subject(email_subject)
                email_subject = email_subject[:300]

                # Make folder name unique if it already exists
                base_email_folder = pdf_folder / email_subject
                email_folder = base_email_folder
                counter = 1
                while email_folder.exists():
                    email_folder = pdf_folder / f"{email_subject}({counter})"
                    counter += 1

                email_folder.mkdir(exist_ok=True)

                log_func(f"\nProcessing email: {email_subject}")

                # Reset city counter for each email
                city_counter.clear()

                processed_count = 0
                manual_review_count = 0

                for attachment in message.Attachments:
                    filename = attachment.FileName

                    # Handle .msg files
                    if filename.lower().endswith('.msg'):
                        log_func(f"\nFound .msg file: {filename}")
                        temp_msg_path = temp_folder / filename
                        attachment.SaveAsFile(str(temp_msg_path))

                        msg_processed, msg_manual, msg_variofix = process_msg_file(
                            temp_msg_path, temp_folder, email_folder, outlook,
                            process_pdf_task2, log_func
                        )
                        processed_count += msg_processed
                        manual_review_count += msg_manual
                        variofix_files.extend(msg_variofix)

                        temp_msg_path.unlink()

                    # Handle PDF files
                    elif filename.lower().endswith('.pdf'):
                        temp_path = temp_folder / filename
                        attachment.SaveAsFile(str(temp_path))
                        log_func(f"\nDownloaded PDF: {filename}")

                        result = process_pdf_task2(
                            temp_path, email_folder, filename)

                        if result['success']:
                            log_func(result['message'])
                            processed_count += 1
                            # Check if message contains dimension warning
                            if 'WARNING' in result['message']:
                                # Extract filename from message "Saved as: CITY 1.pdf ..."
                                saved_filename = result['message'].split(
                                    'Saved as: ')[1].split(' (')[0]
                                files_with_warnings.append(saved_filename)
                        else:
                            log_func(result['message'])
                            manual_review_count += 1
                            # Check if message contains dimension warning
                            if 'WARNING' in result['message']:
                                # Extract filename from message
                                if 'Moved to MANUAL REVIEW:' in result['message']:
                                    saved_filename = result['message'].split(
                                        'Moved to MANUAL REVIEW: ')[1].split(' (')[0]
                                else:
                                    saved_filename = result['message'].split(
                                        'Saved as: ')[1].split(' (')[0]
                                files_with_warnings.append(saved_filename)

                        if result.get('variofix') and result.get('filename'):
                            log_func(f"  ATTENTION VARIOFIX DETECTED in {result['filename']}")
                            variofix_files.append(result['filename'])

                # Create Excel for this email if ANY files were processed OR have dimensions
                if processed_count > 0 or dimensions_data:
                    log_func("\nCreating Excel report for this email...")
                    excel_path = create_excel_report(email_folder)
                    if excel_path:
                        log_func(f"Excel report created: {excel_path.name}")

                log_func(
                    f"Email '{email_subject}': {processed_count} processed, {manual_review_count} need review")
                total_processed_count += processed_count
                total_manual_review_count += manual_review_count

        # Final summary
        log_func("\n" + "="*60)
        log_func("Task 2 Processing complete")
        log_func(f"{total_processed_count} files renamed and saved")
        log_func(f"Location: {pdf_folder}")

        if total_manual_review_count > 0:
            log_func(
                f"\nWARNING: {total_manual_review_count} files need manual review")

        log_func("="*60)

        # Clean up temp folder if empty
        if temp_folder.exists():
            remaining_files = list(temp_folder.glob('*'))
            if not remaining_files:
                temp_folder.rmdir()

        return total_processed_count, total_manual_review_count, pdf_folder, files_with_warnings, variofix_files

    except Exception as e:
        log_func(f"\nError: {e}")
        import traceback
        traceback.print_exc()
        return 0, 0, None, [], []

    finally:
        uninitialize_com()
